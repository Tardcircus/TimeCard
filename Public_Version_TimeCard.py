#!/bin/env python3
import os
import glob
import subprocess
import time
from multiprocessing import Process
from sys import platform
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
import csv
import pandas as pd
import pyexcel as p
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import pwinput
# pip install pyexcel pyexcel-xls pyexcel-xlsx pandas openpyxl


def Title():
        
        os.system('cls||clear')
        print( "")
        print( "       _   _ _   _ascii art of company  ___")
        print( '      |  |   ')
        print( '      |')
        print( '      | | |  ')
        print( '      \ \_|  ')
        print( '       |_/   ')
        print( '   Timecard & Daily Report automation! ')
        print( '' )
        print( '')



def First_Run():
        global path
        global Data_Path

        # Output information
        print("Checking Configuration Files.")

        # Find Program Path
        PWD = Path(__file__).parent
        Data_Path = str(PWD) + "/Data/"
        
        # Add Subdir and pathfile name to Path
        path = str(PWD) + "/Data/config.data"
        path = Path(path)

        #Check if Path file exists. If not then ask for path locations
        if path.is_file():
                pass
        else:
                print(" ")
                print("                 Configuration Error")
                print("         Path locations not configured.")
                print("         If you dont know full paths to Please consult the developer. Bryan@tardcircus.com")
                Configure = input(" Would you like to configure Full paths locations of Timesheet, Dailys, and manpower chart(y,n)? ")
                print(" ")
                
                if Configure != "y":
                        quit()
                else:
                        Config()
        pass


def Config():

        # instructions
        print(" ")
        print(" For windows 10. Right click the folder, select properties, and copy the 'Location'")
        print(" Make sure the path ends with '\'")
        
        # Variables for input of config.data file
        print("")
        User_Name = input("Enter your ******* Email Address: ")
        print(" ")
        Manpower = input("Full/Path/To/Manpower/folder: ")
        print(" ")
        Timecard = input("Full/Path/To/Timecard/folder: ")
        print(" ")
        Daily_Report = input("Full/Path/To/Daily_Report/folder: ")
        print(" ")
        Blank_Docs = input("Full/Path/To/Blank_Daily_and_Timesheet/folder: ")
        print(" ")

        # Make and populate Config file
        file = open(path, "a")
        file.write('{}\n' '{}\n' '{}\n' '{}\n' '{}\n' '{}\n'.format(Manpower, Timecard, Daily_Report, Blank_Docs, Data_Path, User_Name))
        file.close


def Location_Variables():
        
        # Output information
        print("Loading Location Variables.")

        # Apply all folder location variables here!

        # Global access
        global Manpower_Folder
        global Timecard_Folder
        global Daily_Report_Folder
        global Blank_Docs
        global Subdirectory
        global User_Name

        # Access config.data folder and apply variables to data
        Location = open(path).readlines()
        Manpower_Folder = Location[0].strip()
        Timecard_Folder = Location[1].strip()
        Daily_Report_Folder = Location[2].strip()
        Blank_Docs = Location[3].strip()
        Subdirectory = Location[4].strip()
        User_Name = Location[5].strip()


def Make_Data_Columns(Read_File, Print_File, Column_Number):

        # Output information
        print("Making data columns.")

        Data = []
        with open(Read_File, 'r') as f:
                reader = csv.reader(f)

                for row in reader:
                        Data.append(row[Column_Number])
        lines = Data
        file = open(Print_File, 'w')
        for line in lines:
                file.write(f"{line}\n")


def Find_Yesterday():
        global Yesterdays_Day
        global Job_Data_Column1
        global Job_Data_Column2
        global DATE

        # Output information
        print("Finding yesterdays date and day")

        # get yesterdays date of week
        dt = datetime.now()
        dt = dt - timedelta(days=1)
        Yesterdays_Day = dt.strftime('%A')
        DATE = datetime.strftime(dt, '%Y-%m-%d')
        DATE = str(DATE)

        #assign variable to match file column
        if Yesterdays_Day == "Sunday":
                #find Fridays column
                Job_Data_Column1 = 9
                Job_Data_Column2 = 10

        elif Yesterdays_Day == "Monday":
                #find Mondays column
                Job_Data_Column1 = 1
                Job_Data_Column2 = 2

        elif Yesterdays_Day == "Tuesday":
                #find Tuesdays column
                Job_Data_Column1 = 3
                Job_Data_Column2 = 4

        elif Yesterdays_Day == "Wednesday":
                #find Wednesdays column
                Job_Data_Column1 = 5
                Job_Data_Column2 = 6

        elif Yesterdays_Day == "Thursday":
                #find Thursdays column
                Job_Data_Column1 = 7
                Job_Data_Column2 = 8

        elif Yesterdays_Day == "Friday":
                #find Friday column
                Job_Data_Column1 = 9
                Job_Data_Column2 = 10

        elif Yesterdays_Day == "Saturday":
                #find Sridays column
                Job_Data_Column1 = 9
                Job_Data_Column2 = 10

        else:
                print(" ")
                print("         Critical error") 
                print("         Date information not found!")
                print(" ")

def Convert_Data_From_Manpower_Chart():

        # Output info
        print("Running data conversion on Manpower Chart, xls to xlsx format")

        # Find news file in manpower folder
        list_of_files = glob.glob(Manpower_Folder + '*')
        latest_manpower_chart = max(list_of_files, key=os.path.getctime)

        # Check if file is likely a manpower chart and Convert Latest file to xlsx
        if latest_manpower_chart.endswith('.xls'):
                p.save_book_as(file_name=latest_manpower_chart, dest_file_name=Subdirectory + 'Manpower.xlsx')
                print(latest_manpower_chart + " file found, and converted to .xlsx format")

                # Output info
                print("Running data conversion on Manpower chart, xlsx to csv format")

                # Read and store content of an excel file then convert to csv
                read_file = pd.read_excel (Subdirectory + "Manpower.xlsx")
                read_file.to_csv (Subdirectory + "Manpower.csv", index = None, header=True)


def Convert_Data_From_Timesheet():

        # Output info
        print("Running data conversion on Timesheet, xlsx to CSV format")

        # Read and store content of an excel file then convert to csv
        read_file = pd.read_excel (Blank_Docs + 'Blank_timecard.xlsx')
        read_file.to_csv (Subdirectory + "Timesheet.csv", index = None, header=True)


def Pul_Data_From_CSV():
        global Job_List
        global Labor_List1
        global Labor_List2

        # Output information
        print("Pulling data from Manpower chart CSV file")
        
        # Job information Column
        Make_Data_Columns(Subdirectory +'Manpower.csv', Subdirectory + 'Job_Column.data', 0)
        
        # Turn Column File into list
        with open(Subdirectory + 'Job_Column.data') as file:
                Job_List = [line.rstrip() for line in file]

        # Labor Day Column
        Make_Data_Columns(Subdirectory + 'Manpower.csv', Subdirectory + 'Labor_Column1.data', Job_Data_Column1)
        
        # Turn Column File into list
        with open(Subdirectory + 'Labor_Column1.data') as file:
                Labor_List1 = [line.rstrip() for line in file]

        # Labor Day Column 2
        Make_Data_Columns(Subdirectory + 'Manpower.csv', Subdirectory + 'Labor_Column2.data', Job_Data_Column2)
        
        # Turn Column File into list
        with open(Subdirectory + 'Labor_Column2.data') as file:
                Labor_List2 = [line.rstrip() for line in file]


def Make_Daily_Report_Data_File(key):
        global DailyReport_info

        # Output information
        print("Making daily report data file. ")

        # Find the line number of the job in question.
        Job_Number_Location = []
        filename = Subdirectory + 'Job_Column.data'
        with open(filename, 'r') as file:
                lines = file.readlines()
        for number, line in enumerate(lines, 1):
                if key in line:  
                        Job_Number_Location.append(number)
                        print(f'Extracting {key} information from data columns on file line {number}') 

        # Find Job number, and names for DailyReport_info Variable
        Y = len(Job_Number_Location)
        X = 0
        DailyReport_info = []
        DailyReport_info.append(Job_List[Job_Number_Location[X] -1])
        DailyReport_info.append(Job_List[Job_Number_Location[X]])

        while True:
                if X <= Y:
                        DailyReport_info.append(Labor_List1[Job_Number_Location[X] -1])
                        DailyReport_info.append(Labor_List1[Job_Number_Location[X]])
                        DailyReport_info.append(Labor_List2[Job_Number_Location[X] -1])
                        DailyReport_info.append(Labor_List2[Job_Number_Location[X]])
                        X = X + 2
                else:
                        break
        
        # Clean up list and export DailyReport_info to data file
        DailyReport_info = list(filter(None, DailyReport_info))

        for x in range(len(DailyReport_info)):
                DailyReport_info[x] = DailyReport_info[x].strip()
                lines = DailyReport_info
                file = open(Subdirectory + 'DailyReport_info.data', 'w')
                for line in lines:
                        file.write(f"{line}\n")


def Make_Timesheet_Data_File(key):
        global Timesheet_List

        # output information
        print("Making timesheet data file")

        # Pull Employee information Column
        Make_Data_Columns(Subdirectory +'Timesheet.csv', Subdirectory + 'Employee.data', 5)
        
        # Turn Column File into list
        with open(Subdirectory + 'Employee.data') as file:
                Employees_List = [line.rstrip() for line in file]
        R = 131
        Employees_List = Employees_List[R:]
        Employees_List = list(filter(None, Employees_List))

        # Pull true Job information Column
        Make_Data_Columns(Subdirectory +'Timesheet.csv', Subdirectory + 'True_Job.data', 16)
        
        # True Job name to list
        with open(Subdirectory + 'True_Job.data', 'r') as f:
                for line in f.readlines():
                        if key in line:
                                Timesheet_List = [line.rstrip()]

        # Compare Dailyreport manpower with Employee list, and create Timesheet file
        global DailyReport_info
        X = len(Employees_List)
        Y = len(DailyReport_info)

        i = 2
        while True:
                if i < Y:
                        for line in Employees_List:
                                if DailyReport_info[i] in line:
                                        Timesheet_List.append(line)
                        i = i + 1
                else:
                        break

        # Add manpower to list
        Timesheet_List = list(filter(None, Timesheet_List))
        X = len(Timesheet_List)
        Y = 0
        while Y < X:
                print(Timesheet_List[Y])
                Y = Y + 1
        

                # Timecard_Data.data

def Make_Dailyreport():

        # Output information
        print("Building daily report xlsx file")

        JOBINFO = open(Subdirectory + 'DailyReport_info.data').readlines()

        #Import Job information
        JOB = (JOBINFO[0])
        JOBN = JOB.split()[0]
        JOBNAME = JOB.split()[1]
        CREW = JOBINFO
        A = CREW[1]
        B = CREW[0]
        CREW.remove(A)
        CREW.remove(B)
        CREWW = []
        for i in CREW:
                CREWW.append(i.strip())
        CREWWW = ', '.join(CREWW)

        # Daily Report
        workbook = load_workbook(filename=Blank_Docs + "Daily_Report.xlsx")
        sheet = workbook.active
        sheet['C8'] = JOBNAME
        sheet['C10'] = JOBN
        sheet['I8'] = DATE
        sheet['E11'] = CREWWW
        workbook.save(filename=Daily_Report_Folder + "Daily_Report_" + DATE + ".xlsx")


def Make_Timesheet():

        CREWINFO = Timesheet_List # open(Subdirectory + "Timecard_Data.data").readlines()
        JobName = (CREWINFO[0])
        JobName = JobName.strip()

        workbook = load_workbook(filename=Blank_Docs + "Blank_timecard.xlsx")
        sheet = workbook.active
        # Extract Job List

        #BREAK
        JobForeman = CREWINFO[1].strip()
        sheet['C2'] = JobForeman
        JobName = CREWINFO[0].strip()
        M = 0
        F = 7
        X = 1
        NAME = 7
        MAX = len(CREWINFO)
        while X <= (MAX - 1) :
                CrewMember = CREWINFO[X].strip()
                if X <= 7:    
                        sheet['F' + str(F)] = "06:30:00am"
                        F = F + 1
                        sheet['F' + str(F)] = "09:30:00am"
                        F = F + 1
                        sheet['F' + str(F)] = "12:30:00pm"
                        F = F + 1
                        sheet['F' + str(F)] = "01:00:00pm"
                        F = F + 1
                        sheet['F' + str(F)] = "03:00:00pm"
                        F = F + 1
                        sheet['B' + str(NAME)] = CrewMember
                        NAME = NAME + 1
                        sheet['B' + str(NAME)] = JobName
                        NAME = NAME + 4
                X = X + 1

        while X <= F :
                if X <= 7:    
                        sheet['F' + str(F)] = "06:30:00am"
                        F = F + 1
                        sheet['F' + str(F)] = "09:30:00am"
                        F = F + 1
                        sheet['F' + str(F)] = "12:30:00pm"
                        F = F + 1
                        sheet['F' + str(F)] = "01:00:00pm"
                        F = F + 1
                        sheet['F' + str(F)] = "03:00:00pm"
                        F = F + 1
                        NAME = NAME + 1
                        sheet['B' + str(NAME)] = JobName
                        NAME = NAME + 4
                X = X + 1

        sheet['L4'] = DATE

        workbook.save(filename=Timecard_Folder + "Timesheet_" + DATE + ".xlsx")


def Run_Email(Password):
        global Super
        global PM
        global Job
        global Daily_Report
        global Timecard
        global use

        Job = DailyReport_info[0]

        # Find Super
        if 'Dave' in DailyReport_info[1]:
                Super = 'Dave@*******'

        if 'Ray' in DailyReport_info[1]:
                Super = 'Ray@*****'
        
        # Find PM
        if 'Mike' in DailyReport_info[1]:
                Pm = 'Mikes@*******'

        if 'Kyle' in DailyReport_info[1]:
                Pm = 'Kyle*********'

        if 'Wesley' in DailyReport_info[1]:
                Pm = 'Wesley@*****'
        
        if 'Ramon' in DailyReport_info[1]:
                Pm = 'Ramon**********'

            #    Find Latest files for Daily Report
        list_of_files = glob.glob(Daily_Report_Folder + '*')
        Daily_Report = max(list_of_files, key=os.path.getctime)
             #   Find Latest files for TimeSheet
        list_of_files = glob.glob(Timecard_Folder + '*')
        Timecard = max(list_of_files, key=os.path.getctime)


        # Check OS and lauch Daily Report
#        if platform == "linux" or platform == "linux2":
                
                
        subprocess.Popen(["xdg-open", Daily_Report]) #_Folder + "Daily_Report_" + DATE + ".xlsx"])
        subprocess.Popen(["xdg-open", Timecard]) #_Folder + "Timesheet_" + DATE + ".xlsx"])

        print(Daily_Report)
        print(' ')
        use = input("Use internal email system?(y,n) ")
        print(" ")
        print(" ")
        Job = DailyReport_info[0]
        Email_Timesheet(User_Name, Password, Super, Pm, Job)
        Email_Daily_report(User_Name, Password, Super, Pm, Job)
        # linux

#        elif platform == "win32":
#
#                subprocess.call([Daily_Report]) #_Folder + "Daily_Report_" + DATE + ".xlsx"])
#                subprocess.call([Timecard]) #_Folder + "Bryan_Timesheet_" + DATE + ".xlsx"])
#
#                print(" ")
#                Email_Daily_report(User_Name, Password, Super, Pm, Job)
#                Email_Timesheet(User_Name, Password, Super, Pm, Job)
                # Windows...


def Email_Timesheet(User_Name, Password, Super, Pm, Job):
        global use
        
        # Timesheet Variables
        TO_Time = "Dave@*******, " + "Ray@*******"
        CC_Time = Pm + ", Michelle@*******, " + "lilly@*******, " + "Phil*******"
        Sub_Time = "Timesheet, " + DATE

        ################### THIS STATMENT IS JUST FOR DEBUGGING. NO THUNDERBIRD IN FINAL PRODUCT


        if use == "n":

                os.system("thunderbird -compose to=" + TO_Time + ",cc=" + CC_Time + ",subject=" + Sub_Time + ",body=" + Sub_Time + ",attachment=" + Timecard)

        else:
        # Timesheet
                print(" ")
                print(" Processing timesheet email")
                print(" ")
                username = User_Name
                password = Password
                mail_body = Sub_Time
                mail_attachment=Timecard
                mail_attachment_name="Bryan_Timesheet_" + DATE + ".xlsx"

                mimemsg = MIMEMultipart()
                mimemsg['From']=User_Name
                mimemsg['To']=TO_Time
                mimemsg['CC']=CC_Time
                mimemsg['Subject']=Sub_Time
                mimemsg.attach(MIMEText(mail_body, 'plain'))

                with open(mail_attachment, "rb") as attachment:
                        mimefile = MIMEBase('application', 'octet-stream')
                        mimefile.set_payload((attachment).read())
                        encoders.encode_base64(mimefile)
                        mimefile.add_header('Content-Disposition', "attachment; filename= %s" % mail_attachment_name)
                        mimemsg.attach(mimefile)
                        connection = smtplib.SMTP(host='smtp.********.com', port=***)
                        connection.starttls()
                        connection.login(username,password)
                        connection.send_message(mimemsg)
                        connection.quit() 
                        print(" Timesheet Email sent!")


def Email_Daily_report(User_Name, Password, Super, Pm, Job):


        # Daily Report Variables
        TO_Daily = Super + ', ' + Pm
        CC_Daily = "Michelle@*******, *********, Phil@********"
        Sub_Daily = Job + ' - Daily Report ' + DATE


        ################### THIS STATMENT IS JUST FOR DEBUGGING. NO THUNDERBIRD IN FINAL PRODUCT

        if use == "n":
                os.system("thunderbird -compose to=" + TO_Daily + ",cc=" + CC_Daily +",subject=" + Sub_Daily + ",body=" + Sub_Daily + ",attachment=" + Daily_Report)

        else:
        # Daily Report
                username = User_Name
                password = Password
                mail_body = Sub_Daily
                mail_attachment=Daily_Report
                mail_attachment_name="Daily_Report_" + DATE + ".xlsx"

                mimemsg = MIMEMultipart()
                mimemsg['From']=User_Name
                mimemsg['To']=TO_Daily
                mimemsg['CC']=CC_Daily
                mimemsg['Subject']=Sub_Daily
                mimemsg.attach(MIMEText(mail_body, 'plain'))

                with open(mail_attachment, "rb") as attachment:
                        mimefile = MIMEBase('application', 'octet-stream')
                        mimefile.set_payload((attachment).read())
                        encoders.encode_base64(mimefile)
                        mimefile.add_header('Content-Disposition', "attachment; filename= %s" % mail_attachment_name)
                        mimemsg.attach(mimefile)
                        connection = smtplib.SMTP(host='smtp.**********.com', port=***)
                        connection.starttls()
                        connection.login(username,password)
                        connection.send_message(mimemsg)
                        connection.quit() 
                        print(" Daily Report Email sent!")
                        print(" ")
                        print(" ")
                    

def Remove_Data():
        print("  Purging outdated data...")
        os.remove(Subdirectory + 'Manpower.xlsx')
        os.remove(Subdirectory + 'Manpower.csv')
        os.remove(Subdirectory + 'Timesheet.csv')  
        os.remove(Subdirectory + 'Job_Column.data')
        os.remove(Subdirectory + 'Labor_Column1.data')
        os.remove(Subdirectory + 'Labor_Column2.data')
        os.remove(Subdirectory + 'DailyReport_info.data')
        os.remove(Subdirectory + 'Employee.data')
        os.remove(Subdirectory + 'True_Job.data')        

        print("  Purge Complete!")


#-------------------------------------------------------------------------#

 
        # Run title screen
Title()
Password = pwinput.pwinput(prompt='****** Password: ', mask='*')
keyword = input(" Enter yesterdays Job number(No Letters!): ")

        # Run Test for config file and apply variables
First_Run()
Location_Variables()

        # Run Data Building Functions
Find_Yesterday()
Convert_Data_From_Manpower_Chart()
Convert_Data_From_Timesheet()
Pul_Data_From_CSV()
Make_Daily_Report_Data_File(keyword)
Make_Dailyreport()
Make_Timesheet_Data_File(keyword)
Make_Timesheet()
Run_Email(Password)
#Remove_Data()


